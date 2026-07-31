"""
Parses a salary Excel sheet generically, based on its header row (row 1),
rather than hard-coded column letters. This means it keeps working even if
columns are added, removed, or reordered in future months -- as long as an
"EmpId" and "Employee" (name) column exist somewhere in row 1.
"""
import openpyxl


def _clean(v):
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if v is None:
        return ""
    return v


def parse_payroll_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("The uploaded sheet is empty.")

    header_cells = rows[0]
    headers = [str(c.value).strip() if c.value is not None else None for c in header_cells]

    def find_col(options):
        for i, h in enumerate(headers):
            if h and h.strip().lower() in options:
                return i
        return None

    emp_id_idx = find_col({"empid", "employee id", "emp id"})
    name_idx = find_col({"employee", "employee name", "name"})
    dept_idx = find_col({"department"})
    desig_idx = find_col({"designation"})
    net_idx = find_col({"net payable", "net pay"})

    if emp_id_idx is None or name_idx is None:
        raise ValueError(
            "Could not find 'EmpId' and 'Employee' columns in row 1 of the sheet. "
            "Please make sure the header row contains these exact column names."
        )

    ordered_headers = [h for h in headers if h]
    records = []

    for row in rows[1:]:
        values = [c.value for c in row]
        if emp_id_idx >= len(values) or values[emp_id_idx] in (None, ""):
            continue  # skip blank rows

        emp_id = str(_clean(values[emp_id_idx])).strip()

        data = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = values[i] if i < len(values) else None
            data[h] = _clean(v)

        name = str(_clean(values[name_idx])).strip() if name_idx is not None else ""
        dept = str(_clean(values[dept_idx])).strip() if dept_idx is not None else ""
        desig = str(_clean(values[desig_idx])).strip() if desig_idx is not None else ""
        net = str(_clean(values[net_idx])) if net_idx is not None else ""

        records.append(
            {
                "emp_id": emp_id,
                "name": name,
                "department": dept,
                "designation": desig,
                "net_payable": net,
                "data": data,
            }
        )

    if not records:
        raise ValueError("No employee rows were found below the header row.")

    return ordered_headers, records
